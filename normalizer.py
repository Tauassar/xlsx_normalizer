import logging
import os
import re
import time

import colorama
from colorama import Fore
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter

import log
from settings import DEBUG, first_row, first_col, search_strings, fact_sheet_name
from utils import get_export_filename, get_fact_sheet, \
    get_indexes_for_search_values, check_if_sheet_content_positions_satisfies_requirements, \
    check_if_sheet_name_satisfies_requirements


log.setup_logging('DEBUG' if DEBUG else 'INFO')
logger = logging.getLogger(__name__)
file_name = None


def save_to_file(book, file_location):
    """
    Actual save function to save rendered sheets to file
    :param book:object xlsx file needed to save
    :param file_location:str location where file needed to save
    """
    try:
        book.save(file_location)
        logger.info(f'Файл {file_location} успешно сохранен')
    except FileNotFoundError:
        os.mkdir(re.findall("^[a-z, A-z]*/", file_location)[0])
        book.save(file_location)


def write_xlsx(inp_list):
    """
    Write data to the resulting file
    :param inp_list:list list with no duplicates
    :return:list list of data with incorrect IINs
    """

    def set_cell_style(cell):
        cell.border = thin_border
        cell.fill = blueFill
        cell.font = font
        return cell

    def set_header(sheet):
        sheet.cell(row=7, column=7).value = 'В рамках Бизнес-плана'
        sheet.cell(row=7, column=8).value = ''
        sheet.cell(row=6, column=7).value = ''
        sheet.cell(row=6, column=8).value = ''

    try:

        export_to = get_export_filename(file_name)
        book = Workbook()
        sheet = book.active
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        blueFill = PatternFill(start_color='DBE5F1',
                               end_color='DBE5F1',
                               fill_type='solid')
        font = Font(color="1F497D", name='Verdana', size=8)

        i = first_row
        for instance in inp_list:
            col_num = first_col
            for cell_val in instance:
                sheet.cell(row=i, column=col_num).value = cell_val if cell_val is not None else ''
                set_cell_style(sheet.cell(row=i, column=col_num))
                col_num += 1
            i += 1
        sheet.title = 'Факт'
        for idx, col in enumerate(sheet.columns, 1):
            sheet.column_dimensions[get_column_letter(idx)].auto_size = True
        set_header(sheet)
        save_to_file(book, export_to)
    except PermissionError:
        logger.error("Пожалуйста закройте результирующий файл и запустите скрипт заново")


def get_ws_data(ws, ws_indices):
    rows = []

    def get_cell_data(col_name, col_correction, iter_num):
        if ws_indices[col_name] is None:
            return '-'
        return ws[ws_indices[col_name][1] + iter_num][ws_indices[col_name][0] + col_correction].value

    i = 0
    while True:
        row_data = []
        if get_cell_data('Инвестиционный проект/', 0, i) is not None:
            for key in ws_indices.keys():
                row_data.append(get_cell_data(key, 0, i))
                if 'Q' not in key and 'Общий результат' not in key:
                    row_data.append(get_cell_data(key, 1, i))
            rows.append(row_data)
        else:
            print(f'{Fore.WHITE}')
            return rows
        if i % 10 == 0:
            print(f'\r\t\t\t\t{Fore.YELLOW}{i} строк обработано', end='')
        i += 1


def read_xlsx_and_check_if_doc_satisfies_requirements(name):
    """
    Actual save function to save rendered sheets to file
    :param name:name of the xlsx file
    :param index: index, of the particular list in excel doc
    """
    try:
        logger.info("Чтение файла ...")
        wb = load_workbook(name)
        ws = get_fact_sheet(wb)
        logger.info('Проверка входного документа')
        sheet_name_check = check_if_sheet_name_satisfies_requirements(wb)
        sheet_content_check = check_if_sheet_content_positions_satisfies_requirements(ws)
        if sheet_name_check and sheet_content_check:
            logger.info("Проверка пройдена, документ готов для сверки")
            export_to = get_export_filename(file_name)
            save_to_file(wb, export_to)
            return True
        elif not sheet_name_check and sheet_content_check:
            wb.active.title = fact_sheet_name
            export_to = get_export_filename(file_name)
            save_to_file(wb, export_to)
        else:
            logger.info("Индексация столбцов инициированна")
            ws_indices = get_indexes_for_search_values(ws)
            logger.debug(ws_indices)
            logger.info("Индексация столбцов завершена")
            data_rows = get_ws_data(ws, ws_indices)
            logger.info(f'Обработка входного файла завершена, количество обработанных строк: {len(data_rows)}')
            logger.info(f'Запись в выходной файл инициирована')
            write_xlsx(data_rows)
            logger.info(f'Запись в выходной файл завершена, файл находится в папке "out"')
            print(f'{Fore.RED}Необходимо пересохранить файл, для этого откройте файл в Excel и нажммите "Ctrl+S"')
    except ValueError as error:
        logger.error(error)
        logger.info('Выполнение скрипта завершается')


def get_input_file_name():
    logger.info('Поиск Excel файлов в директории')
    files = [f for f in os.listdir('.') if os.path.isfile(f) and (('.xlsx' in f) or ('.xls' in f))]
    if len(files) == 1:
        return files[0]
    else:
        logger.error(f'Удалите лишние Excel файлы из папки и запустите скрипт заново')
        return None


if __name__ == '__main__':
    colorama.init()
    file_name = get_input_file_name()

    if file_name is not None:
        read_xlsx_and_check_if_doc_satisfies_requirements(file_name)
        time.sleep(1)
        print(f'Для выхода нажмите Enter')
        input()
